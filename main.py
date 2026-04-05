import os
import traceback
from collections import OrderedDict
from datetime import datetime
from uuid import uuid4

import pandas as pd
from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
TEMPLATE_ASSET_PATH = os.path.join(ASSETS_DIR, "模版.xlsx")

LOG_FILE = os.path.join(OUTPUT_DIR, "运行日志.txt")

CUSTOMER_SHEET = "到客户"
STORE_SHEET = "到门店"

MAX_PRODUCTS = 4

app = Flask(__name__)


def ensure_dir():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(ASSETS_DIR, exist_ok=True)


def reset_log():
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write("start\n")


def log(msg: str):
    print(msg)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(str(msg) + "\n")


def safe_str(x):
    return "" if pd.isna(x) else str(x).strip()


def norm(x):
    return (
        safe_str(x)
        .replace(" ", "")
        .replace("　", "")
        .replace("\n", "")
        .replace("\r", "")
        .lower()
    )


def parse_products(product_list):
    products = [safe_str(x) for x in product_list if safe_str(x)]
    products = list(dict.fromkeys(products))
    if not products:
        raise ValueError("请至少填写1个产品")
    if len(products) > MAX_PRODUCTS:
        raise ValueError(f"最多支持{MAX_PRODUCTS}个产品")
    return products


def pad_products(products):
    result = products[:]
    while len(result) < MAX_PRODUCTS:
        result.append("")
    return result[:MAX_PRODUCTS]


def detect_column(df: pd.DataFrame, candidates):
    col_map = {norm(c): c for c in df.columns}
    for c in candidates:
        nc = norm(c)
        if nc in col_map:
            return col_map[nc]
    return None


def match_product(name: str, products):
    n = norm(name)
    for p in products:
        if p and norm(p) in n:
            return p
    return None


def load_data(file_path, products):
    df = pd.read_excel(file_path)

    customer_col = detect_column(df, ["所属客户渠道名称", "客户渠道名称", "客户名称", "渠道名称"])
    store_col = detect_column(df, ["门店名称", "门店", "门店名"])
    sn_col = detect_column(df, ["SN", "sn", "sn码", "sn号"])
    product_col = (
        detect_column(df, ["传播名"])
        or detect_column(df, ["产品型号"])
        or detect_column(df, ["Offering别称"])
        or detect_column(df, ["产品名称"])
        or detect_column(df, ["机型名称"])
        or detect_column(df, ["商品名称"])
    )

    if not customer_col:
        raise ValueError("原表缺少客户字段")
    if not store_col:
        raise ValueError("原表缺少门店字段")
    if not sn_col:
        raise ValueError("原表缺少SN字段")
    if not product_col:
        raise ValueError("原表缺少产品字段")

    out = pd.DataFrame()
    out["客户名称"] = df[customer_col].map(lambda x: safe_str(x).split("_")[0])
    out["门店名称"] = df[store_col].map(safe_str)
    out["SN"] = df[sn_col].map(safe_str)
    out["产品名称"] = df[product_col].map(safe_str)
    out["产品归类"] = out["产品名称"].map(lambda x: match_product(x, products))

    out = out[
        (out["产品归类"].notna())
        & (out["客户名称"] != "")
        & (out["门店名称"] != "")
        & (out["SN"] != "")
    ].copy()

    out = out.drop_duplicates(subset=["客户名称", "门店名称", "产品归类", "SN"]).copy()
    return out


def build_customer_rows(df, products):
    ordered_customers = list(OrderedDict.fromkeys(df["客户名称"].tolist()))
    rows = []

    for customer in ordered_customers:
        sub = df[df["客户名称"] == customer]
        row = {"客户名称": customer}
        total = 0
        for p in products:
            qty = int((sub["产品归类"] == p).sum())
            row[p] = qty
            total += qty
        row["合计"] = total
        rows.append(row)

    rows = sorted(rows, key=lambda x: x["合计"], reverse=True)
    return rows


def build_store_rows(df, products):
    ordered_stores = list(OrderedDict.fromkeys(df["门店名称"].tolist()))
    rows = []

    for store in ordered_stores:
        sub = df[df["门店名称"] == store]
        row = {"门店名称": store}
        total = 0
        for p in products:
            qty = int((sub["产品归类"] == p).sum())
            row[p] = qty
            total += qty
        row["合计"] = total
        rows.append(row)

    rows = sorted(rows, key=lambda x: x["合计"], reverse=True)
    return rows


def clear_sheet_data(ws, start_row=2):
    max_row = ws.max_row
    max_col = max(ws.max_column, 6)
    if max_row < start_row:
        return

    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)
            ws.cell(r, c).font = Font(name="微软雅黑", bold=False, color="000000")
            ws.cell(r, c).border = Border()
            ws.cell(r, c).alignment = Alignment(horizontal="general", vertical="center")


def apply_sheet_style(ws, is_store=False):
    """
    到客户:
        A 客户名称 | B-E 产品列 | F 合计
    到门店:
        A 门店名称 | B-E 产品列 | F 合计
    """
    header_fill = PatternFill("solid", fgColor="DCEBFF")
    total_fill = PatternFill("solid", fgColor="EAF3FF")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
    data_font = Font(name="微软雅黑", size=11, bold=False, color="000000")
    total_font = Font(name="微软雅黑", size=11, bold=True, color="000000")

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = 6  # 固定6列：名称 + 4产品 + 合计

    # 行高
    ws.row_dimensions[1].height = 24
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 22

    # 样式
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border

            if r == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                if c == 1:
                    cell.alignment = left_align
                if c == max_col:
                    cell.fill = total_fill
            else:
                cell.font = data_font
                if c == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                if c == max_col:
                    cell.fill = total_fill
                    cell.font = total_font

    # 自动列宽，且四个产品列宽一致
    # A列名称单独算宽度，B-E取统一最大宽度，F合计固定宽度
    def text_width(v):
        s = safe_str(v)
        width = 0
        for ch in s:
            width += 2 if ord(ch) > 127 else 1
        return width

    # A列
    a_width = 10
    for r in range(1, max_row + 1):
        a_width = max(a_width, text_width(ws.cell(r, 1).value))
    ws.column_dimensions["A"].width = min(a_width + 4, 40)

    # B-E 产品列统一宽度
    product_width = 8
    for col_idx in range(2, 6):
        for r in range(1, max_row + 1):
            product_width = max(product_width, text_width(ws.cell(r, col_idx).value))
    product_width = min(product_width + 4, 18)

    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = product_width

    # F 合计
    ws.column_dimensions["F"].width = 10


def write_customer_sheet(ws, rows, products):
    fixed_products = pad_products(products)
    headers = ["客户名称"] + fixed_products + ["合计"]

    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h

    clear_sheet_data(ws, start_row=2)

    for r_idx, row in enumerate(rows, start=2):
        ws.cell(r_idx, 1).value = row["客户名称"]
        col = 2
        row_total = 0

        for p in fixed_products:
            qty = row.get(p, 0) if p else 0
            ws.cell(r_idx, col).value = qty
            row_total += int(qty)
            col += 1

        ws.cell(r_idx, col).value = row_total

    apply_sheet_style(ws, is_store=False)


def write_store_sheet(ws, rows, products):
    fixed_products = pad_products(products)
    headers = ["门店名称"] + fixed_products + ["合计"]

    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h

    clear_sheet_data(ws, start_row=2)

    for r_idx, row in enumerate(rows, start=2):
        ws.cell(r_idx, 1).value = row["门店名称"]
        col = 2
        row_total = 0

        for p in fixed_products:
            qty = row.get(p, 0) if p else 0
            ws.cell(r_idx, col).value = qty
            row_total += int(qty)
            col += 1

        ws.cell(r_idx, col).value = row_total

    apply_sheet_style(ws, is_store=True)


def build_output_name(products):
    product_str = "+".join([safe_str(p).replace("/", "-") for p in products if safe_str(p)])
    product_str = product_str[:40] if product_str else "无产品"
    time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"输出_{product_str}_{time_str}.xlsx"


def process_report(data_file, template_file, products):
    reset_log()
    log(f"本次产品：{products}")
    log(f"模版：{template_file}")
    log(f"数据：{data_file}")

    if not os.path.exists(template_file):
        raise FileNotFoundError(f"未找到模版文件：{template_file}")

    wb = load_workbook(template_file)

    if CUSTOMER_SHEET not in wb.sheetnames:
        raise ValueError(f"模版缺少工作表：{CUSTOMER_SHEET}")
    if STORE_SHEET not in wb.sheetnames:
        raise ValueError(f"模版缺少工作表：{STORE_SHEET}")

    ws_customer = wb[CUSTOMER_SHEET]
    ws_store = wb[STORE_SHEET]

    df = load_data(data_file, products)
    log(f"有效记录数：{len(df)}")

    customer_rows = build_customer_rows(df, products)
    store_rows = build_store_rows(df, products)

    log(f"客户数：{len(customer_rows)}")
    log(f"门店数：{len(store_rows)}")

    write_customer_sheet(ws_customer, customer_rows, products)
    write_store_sheet(ws_store, store_rows, products)

    out_name = build_output_name(products)
    out_path = os.path.join(OUTPUT_DIR, out_name)

    wb.save(out_path)

    log(f"当前输出文件名：{out_name}")
    log("结果Excel输出完成")
    log("全部完成")

    return {
        "customer_rows": customer_rows,
        "store_rows": store_rows,
        "products": products,
        "result_file": out_name,
        "log_file": os.path.basename(LOG_FILE),
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    try:
        ensure_dir()

        data_file = request.files.get("data_file")
        template_file = request.files.get("template_file")

        if not data_file or not template_file:
            return jsonify({"ok": False, "msg": "请上传微服务原表和模版文件"})

        products = parse_products(
            [
                request.form.get("product1", ""),
                request.form.get("product2", ""),
                request.form.get("product3", ""),
                request.form.get("product4", ""),
            ]
        )

        data_name = f"data_{uuid4().hex[:8]}_{data_file.filename}"
        template_name = f"template_{uuid4().hex[:8]}_{template_file.filename}"

        data_path = os.path.join(UPLOAD_DIR, data_name)
        template_path = os.path.join(UPLOAD_DIR, template_name)

        data_file.save(data_path)
        template_file.save(template_path)

        result = process_report(data_path, template_path, products)

        return jsonify(
            {
                "ok": True,
                "msg": "处理完成",
                **result,
            }
        )

    except Exception as e:
        err = str(e)
        tb = traceback.format_exc()
        ensure_dir()
        log(err)
        log(tb)
        return jsonify({"ok": False, "msg": err})


@app.route("/download-template")
def download_template():
    if not os.path.exists(TEMPLATE_ASSET_PATH):
        return "模版文件不存在", 404

    return send_file(
        TEMPLATE_ASSET_PATH,
        as_attachment=True,
        download_name="模版.xlsx",
    )


@app.route("/download/<path:filename>")
def download_file(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


if __name__ == "__main__":
    ensure_dir()
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=True)